"""
CAD file parser and metadata extraction for BRINC app.
"""

import streamlit as st
import pandas as pd
import numpy as np
import os
import re
import io
import json
import datetime
import math
from pathlib import Path

from modules.config import STATE_FIPS, US_STATES_ABBR, KNOWN_POPULATIONS

def _extract_file_meta(raw_df, res_df, filename=""):
    """
    Compute and return a dict of data-matrix statistics from a parsed CAD upload.
    Call this once per file inside aggressive_parse_calls() and store the result
    in st.session_state['file_meta'].  All values are JSON-safe scalars or strings.
    """
    meta = {}
    try:
        meta['uploaded_filename']   = str(filename)
        meta['file_row_count']      = int(len(raw_df))
        meta['file_col_count']      = int(len(raw_df.columns))
        meta['file_col_names']      = json.dumps(list(raw_df.columns))

        # ── City / state inferred from the file ──────────────────────────────
        meta['file_inferred_city']  = str(res_df['_csv_city'].iloc[0])  if '_csv_city'  in res_df.columns and len(res_df) > 0 else ''
        meta['file_inferred_state'] = str(res_df['_csv_state'].iloc[0]) if '_csv_state' in res_df.columns and len(res_df) > 0 else ''

        # ── Date range ───────────────────────────────────────────────────────
        if 'date' in res_df.columns:
            _dates = pd.to_datetime(res_df['date'], errors='coerce').dropna()
            if not _dates.empty:
                meta['file_date_range_start'] = _dates.min().strftime('%Y-%m-%d')
                meta['file_date_range_end']   = _dates.max().strftime('%Y-%m-%d')
                meta['file_date_span_days']   = int((_dates.max() - _dates.min()).days)
                meta['peak_month']            = int(_dates.dt.month.value_counts().idxmax())
                meta['peak_day_of_week']      = int(_dates.dt.dayofweek.value_counts().idxmax())
            else:
                meta['file_date_range_start'] = ''
                meta['file_date_range_end']   = ''
                meta['file_date_span_days']   = 0
                meta['peak_month']            = 0
                meta['peak_day_of_week']      = 0
        else:
            meta['file_date_range_start'] = ''
            meta['file_date_range_end']   = ''
            meta['file_date_span_days']   = 0
            meta['peak_month']            = 0
            meta['peak_day_of_week']      = 0

        # ── Peak hour ────────────────────────────────────────────────────────
        if 'time' in res_df.columns:
            _times = pd.to_datetime(res_df['time'], format='%H:%M:%S', errors='coerce').dropna()
            meta['peak_hour'] = int(_times.dt.hour.value_counts().idxmax()) if not _times.empty else -1
        else:
            meta['peak_hour'] = -1

        # ── Null rate across key CAD fields ──────────────────────────────────
        _key_fields = [c for c in ['lat', 'lon', 'date', 'time', 'priority', 'call_type_desc'] if c in res_df.columns]
        if _key_fields:
            _null_pct = res_df[_key_fields].isnull().values.mean()
            meta['file_null_rate_pct'] = round(float(_null_pct) * 100, 1)
        else:
            meta['file_null_rate_pct'] = 0.0

        # ── Coordinate detection ─────────────────────────────────────────────
        meta['file_has_lat_lon']  = bool('lat' in res_df.columns and 'lon' in res_df.columns and res_df[['lat','lon']].dropna().shape[0] > 0)
        meta['file_has_priority'] = bool('priority' in res_df.columns and res_df['priority'].dropna().shape[0] > 0)

        # ── Call-type breakdown (top 10) ─────────────────────────────────────
        _type_col = next((c for c in ['call_type_desc','agencyeventtypecodedesc','calldesc','description','nature'] if c in res_df.columns), None)
        if _type_col:
            _tc = res_df[_type_col].dropna().str.strip().value_counts().head(10)
            meta['call_type_breakdown'] = json.dumps({str(k): int(v) for k, v in _tc.items()})
        else:
            meta['call_type_breakdown'] = ''

        # ── Priority distribution ─────────────────────────────────────────────
        if 'priority' in res_df.columns:
            _pc = res_df['priority'].dropna().astype(str).value_counts().sort_index()
            meta['priority_distribution'] = json.dumps({str(k): int(v) for k, v in _pc.items()})
        else:
            meta['priority_distribution'] = ''

    except Exception:
        pass
    return meta
def aggressive_parse_calls(uploaded_files):
    all_calls_list = []
    CV = {
        'date': ['received date','incident date','call date','call creation date','calldatetime','call datetime','calltime','timestamp','date','datetime','date time','dispatch date','time received','incdate','date_rept','date_occu','createdtime','created_time','receivedtime','received_time','eventtime','event_time','incidenttime','incident_time','reportedtime','reported_time','entrytime','entry_time','time_central','time_stamp','created'],
        'time': ['call creation time','call time','dispatch time','received time','time', 'hour', 'hour_rept','hour_occu'],
        'priority': ['call priority', 'priority level', 'priority', 'pri', 'urgency'],
        'lat': ['latitude','lat','y coord','ycoord','ycoor','addressy','geoy','y_coord','map_y',
                'point_y','gps_lat','gps_latitude','ylat','coord_y','northing','y_wgs','lat_wgs',
                'incident_lat','inc_lat','event_lat','y_coordinate','address_y','ylocation'],
        'lon': ['longitude','lon','long','x coord','xcoord','xcoor','addressx','geox','x_coord',
                'map_x','point_x','gps_lon','gps_long','gps_longitude','xlon','coord_x','easting',
                'x_wgs','lon_wgs','incident_lon','inc_lon','event_lon','x_coordinate','address_x','xlocation']
    }


    def _looks_like_headerless_geocoder_export(df):
        try:
            cols = [str(c).strip() for c in df.columns]
            coord_pat = re.compile(r'^-?\d+(?:\.\d+)?\s*,\s*-?\d+(?:\.\d+)?$')
            zip_pat = re.compile(r'\b[A-Z]{2}\b\s*,\s*\d{5}(?:-\d{4})?$', re.I)
            has_coord_col = any(coord_pat.match(c) for c in cols)
            has_address_col = any(',' in c and zip_pat.search(c) for c in cols)
            has_matchish = any(str(c).strip().lower() in {'match', 'no_match', 'exact', 'non_exact'} for c in cols)
            return has_coord_col and has_address_col and has_matchish
        except Exception:
            return False

    def _normalize_headerless_geocoder_export(df):
        rows = [list(df.columns)] + df.astype(str).fillna('').values.tolist()
        width = max(len(r) for r in rows)
        padded = [r + [''] * (width - len(r)) for r in rows]
        norm = pd.DataFrame(padded)
        base_cols = ['source_id', 'input_address', 'match_status', 'match_type', 'matched_address', 'lonlat', 'external_id', 'side']
        if width > len(base_cols):
            base_cols += [f'extra_{i}' for i in range(1, width - len(base_cols) + 1)]
        norm.columns = base_cols[:width]
        return norm

    def _extract_lonlat_pair(series):
        s = series.astype(str).str.strip()
        pair = s.str.extract(r'^\s*(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)\s*$')
        lon = pd.to_numeric(pair[0], errors='coerce')
        lat = pd.to_numeric(pair[1], errors='coerce')
        valid = ((lat.between(-90, 90)) & (lon.between(-180, 180))).mean()
        return lon, lat, float(valid)

    def _infer_city_from_location_text(raw_df):
        text_cols = [c for c in raw_df.columns if c in ['location', 'address', 'incident_location', 'addr', 'street', 'input_address', 'matched_address']]
        if not text_cols:
            return None

        s = raw_df[text_cols[0]].dropna().astype(str).str.upper().str.strip()
        if s.empty:
            return None

        s = s.str.replace(r':.*$', '', regex=True)
        s = s.str.replace(r'\bCNTY\b', 'COUNTY', regex=True)
        s = s.str.replace(r'[^A-Z0-9 /,-]', ' ', regex=True)
        s = s.str.replace(r'\s+', ' ', regex=True).str.strip()

        candidates = []
        for val in s:
            padded = f' {val} '
            if ' MOBILE ' in padded:
                candidates.append('Mobile')
                continue

            parts = [p.strip() for p in val.split(',') if p and p.strip()]
            if len(parts) >= 2:
                locality = parts[-2] if re.match(r'^[A-Z]{2}$', parts[-1]) else parts[-1]
                locality = locality.strip()
                if locality and locality not in {'COUNTY', 'CITY', 'TOWN', 'VILLAGE', 'HAMLET'}:
                    candidates.append(locality.title())
                    continue

            m = re.search(r'\b([A-Z]{3,}(?:\s+[A-Z]{3,}){0,2})$', val)
            if m:
                city = m.group(1).title()
                if city not in {'County', 'City', 'Town', 'Village', 'Hamlet'}:
                    candidates.append(city)

        if not candidates:
            return None

        vc = pd.Series(candidates).value_counts()
        return vc.index[0] if not vc.empty else None

    def _infer_state_from_text(raw_df, inferred_city=None):
        for col in ['state', 'state_name']:
            if col in raw_df.columns:
                top = raw_df[col].dropna().astype(str).str.strip().value_counts()
                if not top.empty:
                    state_val = top.index[0]
                    state_up = str(state_val).upper()
                    if state_up in STATE_FIPS:
                        return state_up
                    state_title = str(state_val).title()
                    if state_title in US_STATES_ABBR:
                        return US_STATES_ABBR[state_title]

        if inferred_city == 'Mobile':
            return 'AL'
        return None


    def _choose_priority_column(raw_df):
        exact_names = ['priority', 'call priority', 'priority level', 'pri']
        exact = [c for c in raw_df.columns if c.strip().lower() in exact_names]
        if exact:
            exact.sort(key=lambda c: (
                pd.to_numeric(raw_df[c], errors='coerce').dropna().isin([1,2,3,4,5,6,7,8,9]).mean(),
                -raw_df[c].dropna().nunique()
            ), reverse=True)
            return exact[0]

        loose_names = ['priority', 'call priority', 'priority level', 'pri', 'urgency']
        loose = [c for c in raw_df.columns if any(k in c for k in loose_names)]
        if loose:
            loose.sort(key=lambda c: (
                pd.to_numeric(raw_df[c], errors='coerce').dropna().isin([1,2,3,4,5,6,7,8,9]).mean(),
                -raw_df[c].dropna().nunique()
            ), reverse=True)
            return loose[0]
        return None

    def parse_priority(raw):
        s = str(raw).strip().upper()
        if not s or s == 'NAN': return None
        # Smart inference for PD offenses if priority column is missing
        if any(w in s for w in ['ROBBERY','BURGLARY','ASSAULT','SHOOTING','STABBING','CRITICAL','EMERG']): return 1
        if any(w in s for w in ['ACCIDENT','DISTURBANCE','THEFT','MED','ALARM']): return 2
        if any(w in s for w in ['NON REPORTABLE','FOUND PROPERTY','INFO','ROUTINE','MISC']): return 4
        
        m = re.search(r'^(\d+)', s)
        if m: return int(m.group(1))
        return 3

    for cfile in uploaded_files:
        try:
            fname = cfile.name.lower()
            excel_exts = ('.xlsx', '.xls', '.xlsb', '.xlsm')

            if fname.endswith(excel_exts):
                # ── Excel path ────────────────────────────────────────────────
                raw_bytes = cfile.getvalue()
                engine = 'openpyxl'
                if fname.endswith('.xls'):
                    engine = 'xlrd'
                elif fname.endswith('.xlsb'):
                    engine = 'pyxlsb'

                def _sheet_score(ws):
                    score = 0
                    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
                    if not rows:
                        return -1
                    header = rows[0] or []
                    header_norm = [str(h).strip().lower() for h in header if h is not None]
                    if not header_norm:
                        return -1
                    hints = ['latitude', 'longitude', 'lat', 'lon', 'priority', 'location', 'date', 'time']
                    score += sum(10 for h in header_norm if any(k == h or k in h for k in hints))
                    score += sum(1 for h in header_norm if h and not re.match(r'^column\d+$', h))
                    if len(rows) > 1 and rows[1] and any(v is not None and str(v).strip() != '' for v in rows[1]):
                        score += 25
                    # Penalize external-data placeholder sheets
                    if len(header_norm) == 1 and header_norm[0].startswith('externaldata_'):
                        score -= 100
                    return score

                try:
                    import openpyxl as _oxl
                    _wb = _oxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
                    _sheet_name = max(_wb.sheetnames, key=lambda sn: _sheet_score(_wb[sn]))
                    _ws = _wb[_sheet_name]
                    _row_iter = _ws.iter_rows(values_only=True)
                    _headers_raw = next(_row_iter)
                    if _headers_raw is None:
                        raise ValueError("Selected Excel sheet has no header row.")
                    _real_idx = [
                        i for i, h in enumerate(_headers_raw)
                        if h is not None and not (str(h).startswith('Column') and str(h)[6:].isdigit())
                    ]
                    if not _real_idx:
                        _real_idx = [i for i, h in enumerate(_headers_raw) if h is not None]
                    _real_headers = [str(_headers_raw[i]).lower().strip() for i in _real_idx]
                    _rows_data = []
                    for _row in _row_iter:
                        if _row is None:
                            continue
                        _trimmed = [_row[i] if i < len(_row) else None for i in _real_idx]
                        if any(v is not None and str(v).strip() != '' for v in _trimmed):
                            _rows_data.append(_trimmed)
                    _wb.close()
                    raw_df = pd.DataFrame(_rows_data, columns=_real_headers)
                    raw_df = raw_df.dropna(how='all')
                    raw_df.columns = [str(c).lower().strip() for c in raw_df.columns]
                except Exception as _xe:
                    raw_df = None
                    # Try all sheets with pandas and pick the one that looks most like CAD data
                    try:
                        _all = pd.read_excel(io.BytesIO(raw_bytes), engine=engine, sheet_name=None)
                        best_score = -10**9
                        best_df = None
                        for _sn, _df in _all.items():
                            _df.columns = [str(c).lower().strip() for c in _df.columns]
                            _score = 0
                            for _c in _df.columns:
                                if _c in ('latitude', 'longitude', 'priority', 'location'):
                                    _score += 20
                                elif any(k in _c for k in ['lat', 'lon', 'priority', 'location', 'date', 'time']):
                                    _score += 5
                            _score += min(len(_df), 100)
                            if len(_df.columns) == 1 and str(_df.columns[0]).startswith('externaldata_'):
                                _score -= 100
                            if _score > best_score:
                                best_score = _score
                                best_df = _df
                        if best_df is not None:
                            raw_df = best_df
                    except Exception:
                        pass
                    if raw_df is None:
                        raw_df = pd.read_excel(io.BytesIO(raw_bytes), engine=engine, dtype=str)
                        raw_df.columns = [str(c).lower().strip() for c in raw_df.columns]
            else:
                # ── CSV / TXT path ────────────────────────────────────────────
                content = cfile.getvalue().decode('utf-8', errors='ignore')
                first_line = content.split('\n')[0]
                delim = ',' if first_line.count(',') > first_line.count('\t') else '\t'
                raw_df = pd.read_csv(io.StringIO(content), sep=delim, dtype=str)
                if _looks_like_headerless_geocoder_export(raw_df):
                    raw_df = _normalize_headerless_geocoder_export(raw_df)
                raw_df.columns = [str(c).lower().strip() for c in raw_df.columns]

                # ── Census Geocoder: split combined 'lonlat' column ──────────
                # After normalization the file has a 'lonlat' column storing
                # "lon,lat" pairs (e.g. "-93.283,36.601").  The generic
                # coord-name scanner below matches 'lonlat' for BOTH lat and
                # lon but pd.to_numeric returns all-NaN on comma-pair strings.
                # Handle it here explicitly before the scanner runs, then drop
                # No_Match rows whose lonlat is empty.
                if 'lonlat' in raw_df.columns and 'lat' not in raw_df.columns:
                    _pair = raw_df['lonlat'].astype(str).str.strip().str.extract(
                        r'^\s*(-?[\d.]+)\s*,\s*(-?[\d.]+)\s*$'
                    )
                    _lon_cand = pd.to_numeric(_pair[0], errors='coerce')
                    _lat_cand = pd.to_numeric(_pair[1], errors='coerce')
                    # Census geocoder stores lon first, lat second — verify US range
                    if _lon_cand.between(-180, -50).mean() > 0.3 and _lat_cand.between(18, 72).mean() > 0.3:
                        raw_df['lon'] = _lon_cand
                        raw_df['lat'] = _lat_cand
                    # Drop rows with no geocoded location (No_Match rows)
                    if 'lat' in raw_df.columns:
                        raw_df = raw_df[raw_df['lat'].notna()].copy()

                # ── Census Geocoder: extract city & state from matched_address ──
                # matched_address format: "32 GOLFSHORES DR, BRANSON, MO, 65616"
                if 'matched_address' in raw_df.columns and '_csv_city' not in raw_df.columns:
                    try:
                        _ma = raw_df['matched_address'].dropna().astype(str)
                        _ma_parts = _ma.str.split(',')
                        _ma_cities = _ma_parts.apply(
                            lambda p: p[-3].strip().title() if len(p) >= 4 else (p[-2].strip().title() if len(p) >= 2 else None)
                        ).dropna()
                        _ma_states = _ma_parts.apply(
                            lambda p: p[-2].strip().upper() if len(p) >= 2 else None
                        ).dropna()
                        if not _ma_cities.empty:
                            raw_df['_csv_city'] = _ma_cities.value_counts().index[0]
                        if not _ma_states.empty:
                            _top_st = _ma_states.value_counts().index[0]
                            if _top_st in STATE_FIPS:
                                raw_df['_csv_state'] = _top_st
                    except Exception:
                        pass

            res = pd.DataFrame()
            exact_coord_names = {
                'lat': ['latitude', 'lat', 'gps_lat', 'gps_latitude'],
                'lon': ['longitude', 'lon', 'long', 'gps_lon', 'gps_longitude']
            }
            for field in ['lat', 'lon']:
                found_exact = [c for c in raw_df.columns if c.strip().lower() in exact_coord_names[field]]
                # Exclude bare 'lonlat' from the loose scan — it's a combined field,
                # not a plain numeric column, and will produce all-NaN via pd.to_numeric.
                found_loose = [c for c in raw_df.columns
                               if c != 'lonlat' and any(s in c for s in CV[field])]
                found = found_exact or found_loose
                if found:
                    res[field] = pd.to_numeric(raw_df[found[0]], errors='coerce')

            if 'lat' not in res.columns or 'lon' not in res.columns:
                for c in raw_df.columns:
                    lon_series, lat_series, valid_rate = _extract_lonlat_pair(raw_df[c])
                    if valid_rate >= 0.50:
                        res['lon'] = lon_series
                        res['lat'] = lat_series
                        break

            # ── Fallback: no column name matched — scan numeric columns by value range ──
            # Lat: -90 to 90, Lon: -180 to 180. Pick best candidate for each.
            if 'lat' not in res.columns or 'lon' not in res.columns:
                numeric_cols = []
                for c in raw_df.columns:
                    series = pd.to_numeric(raw_df[c], errors='coerce').dropna()
                    if len(series) > 10:
                        numeric_cols.append((c, series))

                lat_candidates, lon_candidates = [], []
                for c, series in numeric_cols:
                    mn, mx = series.min(), series.max()
                    # Reject if already assigned
                    if c in (res.get('_lat_col',''), res.get('_lon_col','')):
                        continue
                    # Large integer coords (State Plane) — treat as potential coord pair
                    if mx > 1000:
                        lat_candidates.append((c, series))
                        lon_candidates.append((c, series))
                        continue
                    if -90 <= mn and mx <= 90 and mn < -1:
                        lat_candidates.append((c, series))
                    if -180 <= mn and mx <= 180 and (mn < -90 or mx > 90):
                        lon_candidates.append((c, series))

                # Prefer candidate whose name hints at lat/lon
                def _score(name, hints):
                    return sum(1 for h in hints if h in name)

                if 'lat' not in res.columns and lat_candidates:
                    lat_candidates.sort(key=lambda x: -_score(x[0], ['lat','y','north']))
                    best_lat_col = lat_candidates[0][0]
                    res['lat'] = pd.to_numeric(raw_df[best_lat_col], errors='coerce')

                if 'lon' not in res.columns and lon_candidates:
                    # Don't reuse the lat column
                    used = res.get('lat', pd.Series()).name if 'lat' in res.columns else None
                    lon_candidates = [(c, s) for c, s in lon_candidates if c != used]
                    if lon_candidates:
                        lon_candidates.sort(key=lambda x: -_score(x[0], ['lon','long','x','east']))
                        best_lon_col = lon_candidates[0][0]
                        res['lon'] = pd.to_numeric(raw_df[best_lon_col], errors='coerce')
            
            _p_col = _choose_priority_column(raw_df)
            p_found = [_p_col] if _p_col else []
            if _p_col:
                parsed_priority = raw_df[_p_col].apply(parse_priority)
                parsed_priority = pd.to_numeric(parsed_priority, errors='coerce')
                parsed_priority = parsed_priority.where(parsed_priority.isin([1, 2, 3, 4, 5, 6, 7, 8, 9]))
                if parsed_priority.dropna().empty:
                    res['priority'] = 3
                else:
                    res['priority'] = parsed_priority.fillna(3).astype(int)
            else:
                # No trustworthy priority field — keep the app usable with a neutral default
                res['priority'] = 3
            
            # ── Event type description — carried through for CAD analytics charts ──
            _desc_hints = ['desc','type','nature','offense','calltype','call_type','event_type',
                           'eventtype','calldesc','incident_type','agencyeventtype','violation','call_nature','cfs_type']
            _desc_found = [c for c in raw_df.columns
                           if any(h in c for h in _desc_hints)
                           and c not in (p_found[:1] if p_found else [])]
            if _desc_found:
                # Pick the column with the most unique text values (most descriptive)
                _best_desc = max(_desc_found, key=lambda c: raw_df[c].dropna().nunique())
                if raw_df[_best_desc].dropna().nunique() > 2:
                    res['call_type_desc'] = raw_df[_best_desc].astype(str).str.strip()

            d_found = [c for c in raw_df.columns if any(s in c for s in CV['date'])]
            t_found = [c for c in raw_df.columns if any(s in c for s in CV['time'])]

            # Fallback: if no date column found by name hint, scan all string columns
            # for any that successfully parse as datetime (catches columns like
            # 'createdtime_central', 'call_ts', 'event_dttm', etc.)
            if not d_found:
                for _col in raw_df.columns:
                    if _col in (t_found or []):
                        continue
                    try:
                        _test = pd.to_datetime(raw_df[_col].dropna().head(50), errors='coerce')
                        _valid = _test.dropna()
                        if len(_valid) >= 10 and _valid.dt.year.between(2000, 2035).mean() > 0.8:
                            d_found = [_col]
                            break
                    except Exception:
                        continue

            if d_found:
                # ── Normalise date/time columns to plain strings ───────────────
                # openpyxl returns Python datetime.datetime / datetime.time objects
                # for date and time cells.  Concatenating them with strings via
                # `+ ' '` raises TypeError and is caught by the bare except,
                # silently discarding the entire file.  Convert to strings first.
                def _col_to_datestr(series):
                    """Convert a column that may contain datetime objects → 'YYYY-MM-DD' strings."""
                    try:
                        _p = pd.to_datetime(series, errors='coerce')
                        if _p.notna().mean() > 0.6:
                            return _p.dt.strftime('%Y-%m-%d').where(_p.notna(), '')
                    except Exception:
                        pass
                    return series.fillna('').astype(str).str.strip()

                def _col_to_timestr(series):
                    """Convert a column that may contain datetime.time objects → 'HH:MM:SS' strings."""
                    try:
                        _first = series.dropna().iloc[0] if not series.dropna().empty else None
                        if _first is not None and hasattr(_first, 'strftime'):
                            # openpyxl datetime.time or datetime.datetime objects
                            return series.apply(
                                lambda v: v.strftime('%H:%M:%S') if hasattr(v, 'strftime') else ''
                            ).fillna('')
                    except Exception:
                        pass
                    return series.fillna('').astype(str).str.strip()

                # Build the raw string series to parse — combine date+time cols if separate
                if t_found and d_found[0] != t_found[0]:
                    _raw_dt_str = _col_to_datestr(raw_df[d_found[0]]) + ' ' + _col_to_timestr(raw_df[t_found[0]])
                else:
                    _raw_dt_str = raw_df[d_found[0]].fillna('').astype(str).str.strip()

                # Try explicit common formats first (orders of magnitude faster than
                # dateutil fallback on large files, and avoids NaT on ghost rows).
                # Format detection: sample the first non-null value.
                _sample_vals = _raw_dt_str.dropna().astype(str).str.strip()
                _sample_vals = _sample_vals[_sample_vals != ''].head(5)
                _fmt_candidates = [
                    '%m/%d/%Y %I:%M %p',   # 2/14/2025 6:03 PM  (Mobile AL)
                    '%m/%d/%Y %H:%M:%S',   # 2/14/2025 18:03:00
                    '%m/%d/%Y %H:%M',      # 2/14/2025 18:03
                    '%Y-%m-%d %H:%M:%S',   # 2025-02-14 18:03:00
                    '%Y-%m-%dT%H:%M:%S',   # ISO 8601
                    '%Y-%m-%d %H:%M',      # 2025-02-14 18:03
                    '%Y/%m/%d %H:%M:%S',
                    '%d/%m/%Y %H:%M:%S',
                    '%m-%d-%Y %H:%M:%S',
                ]
                dt_series = None
                if not _sample_vals.empty:
                    for _fmt in _fmt_candidates:
                        try:
                            _trial = pd.to_datetime(_sample_vals.iloc[0], format=_fmt, errors='raise')
                            # Format matched — apply to full series
                            dt_series = pd.to_datetime(_raw_dt_str, format=_fmt, errors='coerce')
                            break
                        except Exception:
                            continue
                if dt_series is None:
                    # Final fallback: let pandas infer (slow but handles edge cases)
                    dt_series = pd.to_datetime(_raw_dt_str, errors='coerce')

                res['date'] = dt_series.dt.strftime('%Y-%m-%d')
                res['time'] = dt_series.dt.strftime('%H:%M:%S')

            # --- COORDINATE CLEANUP: sentinel values & sign errors ---
            if not res.empty and 'lat' in res.columns and 'lon' in res.columns:
                # Drop obvious sentinel/null-coordinate rows before any further processing
                # lat=0 and lon=0 are common CAD null sentinels (no valid location on the equator/prime meridian)
                # lon=-179.99999 is another common sentinel used by some CAD vendors
                _sentinel_mask = (
                    (res['lat'] == 0) | (res['lon'] == 0) |
                    (res['lat'].abs() < 0.001) | (res['lon'].abs() < 0.001) |
                    (res['lon'] < -179.9)
                )
                if _sentinel_mask.any():
                    res = res[~_sentinel_mask].copy()

                # Fix wrong-sign longitudes: some CAD exports omit the minus sign for
                # western-hemisphere longitudes (e.g. 81.31 instead of -81.31).
                # Detect by checking if the majority of lons are negative (correct for US)
                # while a small minority are positive with the same absolute magnitude.
                if not res.empty and 'lon' in res.columns:
                    _neg_count = (res['lon'] < 0).sum()
                    _pos_count = (res['lon'] > 0).sum()
                    _total = len(res)
                    # If >90% are negative but some are positive AND the median negative lon
                    # matches -(positive lon range), flip the positive ones
                    if _neg_count > 0 and _pos_count > 0 and (_neg_count / _total) > 0.90:
                        _median_neg = res.loc[res['lon'] < 0, 'lon'].median()
                        _pos_vals = res.loc[res['lon'] > 0, 'lon']
                        # Check if flipping would land near the median negative cluster
                        _would_match = ((-_pos_vals).between(_median_neg - 2, _median_neg + 2)).mean()
                        if _would_match > 0.5:
                            res.loc[res['lon'] > 0, 'lon'] = -res.loc[res['lon'] > 0, 'lon']

            # --- COORDINATE CONVERSION (STATE PLANE / LARGE-INTEGER DETECTOR) ---
            if not res.empty and 'lat' in res.columns and 'lon' in res.columns:
                res = res[(res['lat'] != 0) & (res['lon'] != 0)].dropna(subset=['lat', 'lon'])
                if not res.empty:
                    max_val = max(res['lat'].abs().max(), res['lon'].abs().max())
                    if max_val > 1000:
                        converted = False
                        # Strategy 1: Try common State Plane CRS at /100 and /1 scales
                        candidate_crs = [
                            "EPSG:2278",  # TX South Central (ftUS)
                            "EPSG:2277",  # TX Central (ftUS)
                            "EPSG:2276",  # TX North Central (ftUS)
                            "EPSG:2279",  # TX South (ftUS)
                            "EPSG:32140", # TX South Central (m)
                        ]
                        for scale in [100.0, 1.0]:
                            for crs in candidate_crs:
                                try:
                                    transformer = pyproj.Transformer.from_crs(crs, "EPSG:4326", always_xy=True)
                                    test_lons, test_lats = transformer.transform(
                                        res['lon'].values[:20] / scale,
                                        res['lat'].values[:20] / scale
                                    )
                                    if (24 < float(test_lats.mean()) < 50 and
                                            -130 < float(test_lons.mean()) < -60 and
                                            float(test_lats.std()) < 5 and
                                            float(test_lons.std()) < 5):
                                        lons, lats = transformer.transform(
                                            res['lon'].values / scale, res['lat'].values / scale
                                        )
                                        res['lon'], res['lat'] = lons, lats
                                        converted = True
                                        break
                                except Exception:
                                    continue
                            if converted:
                                break

                        # Strategy 2: If CRS conversion failed, anchor to city column geocode
                        if not converted:
                            try:
                                city_name = None
                                for col in ['city', 'city_name', 'municipality', 'jurisdiction']:
                                    if col in raw_df.columns:
                                        top = raw_df[col].dropna().str.strip().value_counts()
                                        if not top.empty:
                                            city_name = top.index[0]
                                            break
                                state_name = None
                                for col in ['state', 'state_name']:
                                    if col in raw_df.columns:
                                        top = raw_df[col].dropna().str.strip().value_counts()
                                        if not top.empty:
                                            state_name = top.index[0]
                                            break
                                if city_name:
                                    query_str = f"{city_name}, {state_name}" if state_name else city_name
                                    geo_url = f"https://nominatim.openstreetmap.org/search?format=json&q={urllib.parse.quote(query_str)}&limit=1"
                                    req = urllib.request.Request(geo_url, headers={"User-Agent": "BRINC_COS_Optimizer/1.0"})
                                    with urllib.request.urlopen(req, timeout=10) as resp:
                                        geo_data = json.loads(resp.read().decode("utf-8"))
                                    if geo_data:
                                        anchor_lat = float(geo_data[0]["lat"])
                                        anchor_lon = float(geo_data[0]["lon"])
                                        raw_cx = res["lon"].median()
                                        raw_cy = res["lat"].median()
                                        city_radius_deg = 0.35
                                        raw_spread = max(res["lon"].std(), res["lat"].std(), 1)
                                        deg_per_unit = city_radius_deg / raw_spread
                                        res["lon"] = anchor_lon + (res["lon"] - raw_cx) * deg_per_unit
                                        res["lat"] = anchor_lat + (res["lat"] - raw_cy) * deg_per_unit
                                        converted = True
                            except Exception:
                                pass

                        if converted:
                            res = res[
                                (res["lat"] > 18) & (res["lat"] < 72) &
                                (res["lon"] > -170) & (res["lon"] < -60)
                            ]

            # ── Agency / source tagging (Fire vs Police) ─────────────────────
            # Look for a column named 'agency', 'department', or 'dept' and
            # carry it through as a lowercase 'agency' column so the map
            # renderer can colour fire calls red and police calls the default colour.
            _agency_col = next(
                (c for c in raw_df.columns if c.strip().lower() in ('agency', 'department', 'dept')),
                None
            )
            if _agency_col:
                res['agency'] = raw_df[_agency_col].astype(str).str.strip().str.lower()
            else:
                res['agency'] = 'police'   # safe default for single-agency files

            # Agency / source tagging (Fire vs Police)
            # Prefer a column that contains fire/police labels when duplicate agency
            # fields exist (for example Agency + Agency.1 after CSV import).
            _agency_candidates = [
                c for c in raw_df.columns
                if c.strip().lower() in ('agency', 'agency.1', 'department', 'department.1', 'dept', 'dept.1')
            ]
            _agency_col = None
            for _cand in _agency_candidates:
                try:
                    _vals = raw_df[_cand].astype(str).str.strip().str.lower()
                    if _vals.str.contains(r'\b(fire|police|ems|sheriff)\b', regex=True, na=False).any():
                        _agency_col = _cand
                        break
                except Exception:
                    pass
            if _agency_col is None:
                _agency_col = _agency_candidates[0] if _agency_candidates else None

            if _agency_col:
                res['agency'] = raw_df[_agency_col].astype(str).str.strip().str.lower()
            else:
                res['agency'] = 'police'   # safe default for single-agency files

            # City/state detection: store top values on rows for location detection
            top_city_name = None
            for col in ["city", "city_name", "municipality", "jurisdiction"]:
                if col in raw_df.columns:
                    top_city = raw_df[col].dropna().astype(str).str.strip().value_counts()
                    if not top_city.empty:
                        top_city_name = str(top_city.index[0]).title()
                        res["_csv_city"] = top_city_name
                        break

            if "_csv_city" not in res.columns:
                inferred_city = _infer_city_from_location_text(raw_df)
                if inferred_city:
                    top_city_name = inferred_city
                    res["_csv_city"] = inferred_city

            inferred_state = _infer_state_from_text(raw_df, top_city_name)
            if not inferred_state:
                for _addr_col in ['input_address', 'matched_address', 'address', 'location']:
                    if _addr_col in raw_df.columns:
                        _addr_series = raw_df[_addr_col].astype(str)
                        # Pattern 1: "..., ST, ZIPCODE" (comma-separated state and zip)
                        _states = _addr_series.str.extract(r',\s*([A-Z]{2})\s*,\s*\d{5}(?:-\d{4})?')[0].dropna()
                        if _states.empty:
                            # Pattern 2: "..., ST ZIPCODE" (state and zip in same segment, common format)
                            _states = _addr_series.str.extract(r',\s*([A-Z]{2})\s+\d{5}(?:-\d{4})?')[0].dropna()
                        if not _states.empty:
                            inferred_state = _states.value_counts().idxmax()
                            break
            if inferred_state:
                res["_csv_state"] = inferred_state

            # ── Capture file data matrix for Sheets/email logging ────────────
            try:
                _meta = _extract_file_meta(raw_df, res, filename=cfile.name)
                # Merge into session-level file_meta (last file wins for per-field values;
                # accumulate filenames if multiple files are uploaded at once)
                _existing = st.session_state.get('file_meta', {})
                _existing_names = _existing.get('uploaded_filename', '')
                if _existing_names and _meta.get('uploaded_filename','') and _meta['uploaded_filename'] not in _existing_names:
                    _meta['uploaded_filename'] = _existing_names + ' | ' + _meta['uploaded_filename']
                st.session_state['file_meta'] = {**_existing, **_meta}
            except Exception:
                pass

            all_calls_list.append(res)
        except: continue
        
    if not all_calls_list: return pd.DataFrame()
    # Only keep frames that actually have lat/lon columns — Excel sheets
    # without coordinate data should not crash the concat
    valid = [df for df in all_calls_list if 'lat' in df.columns and 'lon' in df.columns]
    if not valid: return pd.DataFrame()
    combined = pd.concat(valid, ignore_index=True)
    # Safe dropna — columns guaranteed to exist now
    combined = combined.dropna(subset=['lat', 'lon'])
    combined['lat'] = pd.to_numeric(combined['lat'], errors='coerce')
    combined['lon'] = pd.to_numeric(combined['lon'], errors='coerce')
    combined = combined[(combined['lat'].between(-90, 90)) & (combined['lon'].between(-180, 180))]
    # IMPORTANT: keep the full parsed CAD dataset here.
    #
    # The optimizer is sampled later (after upload) for performance, but the
    # parsed dataframe itself must preserve every incident so:
    #   1) Total Incidents shows the true uploaded count
    #   2) the stations map can render a much denser full-history call cloud
    #   3) export/reporting math stays tied to the source file, not a k-means
    #      surrogate created during parsing
    return combined

def _get_annualized_calls(raw_count: int) -> int:
    """Return raw_count scaled to a full year using the uploaded file's date span.

    If the file covers less than a full year (and at least 14 days), the raw
    count is extrapolated to 365 days.  Falls back to raw_count when no date
    span is available (simulated data, unknown span, or span ≥ 330 days).
    """
    span_days = int(st.session_state.get('file_meta', {}).get('file_date_span_days', 0) or 0)
    if 14 <= span_days < 330:
        return round(raw_count * 365 / span_days)
    return raw_count


def _build_apprehension_table(df_calls, text_main, text_muted, card_bg, card_border, accent_color):
